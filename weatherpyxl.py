#importing Required Packages
import requests
from openpyxl import load_workbook
from openpyxl import Workbook

workbook_name=r"C:\Users\ar28s\Documents\openpyxlproject\weatherupdate.xlsx"
wb=load_workbook(workbook_name)
page=wb.active

#Getting the data from OpenWeatherMap using the API
def weather_data(query):

    res=requests.get('http://api.openweathermap.org/data/2.5/weather?'+query+'&APPID=b35975e18dc93725acb092f7272cc6b8&units=metric')
    return res.json()


def print_weather(result,city):
    return [city,result['main']['temp'],result['main']['humidity']]
    

#Specifying City Name
city=input('Enter the city:')

        
try:

        query='q='+city
        w_data=weather_data(query)

        y=print_weather(w_data, city)




        x=[]
        for row in page.rows:
            x.append(row[0].value)
        if city in x[1:]:
            for i in range(2, page.max_row+1):
                cell=page.cell(row=i,column=1)
                up=page.cell(row=i,column=5)
                weather=page.cell(row=i,column=2)
                humdity=page.cell(row=i,column=3)
                typee=page.cell(row=i,column=4)
                if cell.value==city and up.value==1:
                        print("Already available and no updation allowed")
                
                if cell.value==city and up.value==0:
                        temptype=input('C or f:')


                        if temptype=='F':
                            y[1]=(y[1]*(9/5))+32
                        else:
                            y[1]=y[1]
                        page.delete_rows(i)
                        u=int(input("To stop updating further press 1 or else 0:"))
                        z=[cell.value,weather.value,humdity.value,temptype,u]
                        print(z)
                        page.append(z)
                        break
                

        else:
                temptype=input('C or f:')
                if temptype=='F':
                    y[1]=(y[1]*(9/5))+32
                else:
                    y[1]=y[1]
                updation=int(input('Need updation or not,Click 1 to stop or 0:'))
                y+=[temptype,updation]
                page.append(y)
                print("Value Entered Successfully")
        wb.save(filename=workbook_name)
except:
        print("No such city")




    
